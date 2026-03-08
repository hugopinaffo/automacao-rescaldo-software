import os
import shutil
import tempfile
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter, range_boundaries

# ==================================================
# CONFIGURAÇÃO DE COLUNAS E TABELAS
# ==================================================
PAI_COL_NUMERO, PAI_COL_WO, PAI_COL_DESCRICAO = 1, 2, 3
ATU_COL_CHAMADO_PAI, ATU_COL_HOSTNAME, ATU_COL_DESCRICAO, ATU_COL_STATUS = 1, 2, 4, 5

NOME_TABELA_PAI = "Chamado_Pai"
NOME_TABELA_ATUACAO = "Atuação"


# ==================================================
# EXCEÇÃO CUSTOMIZADA
# ==================================================
class AutomacaoErro(Exception):
    pass


def abortar(msg):
    raise AutomacaoErro(msg)


# ==================================================
# UTILITÁRIOS
# ==================================================
def gerar_descricao(software, cve):
    return f"Atualização {software}, referente à {cve}"


def validar_campos(*campos):
    if not all(str(c).strip() if c else False for c in campos):
        abortar("Todos os campos são obrigatórios.")


def validar_extensao_excel(caminho):
    if not caminho.lower().endswith((".xlsx", ".xlsm")):
        abortar("O arquivo de hostnames deve ser .xlsx ou .xlsm.")


def validar_arquivo(caminho, descricao):
    if not os.path.exists(caminho):
        abortar(f"{descricao} não encontrado.")


def obter_tabela(ws, nome):
    nome_lower = nome.lower()
    tabela = next((ws.tables[t] for t in ws.tables if t.lower() == nome_lower), None)
    if not tabela:
        abortar(f"Tabela '{nome}' não encontrada na aba '{ws.title}'.")
    return tabela


def expandir_tabela(ws, nome_tabela, linhas=1):
    tabela = obter_tabela(ws, nome_tabela)
    min_c, min_r, max_c, max_r = range_boundaries(tabela.ref)
    novo_max_r = max_r + linhas

    col_min, col_max = get_column_letter(min_c), get_column_letter(max_c)
    tabela.ref = f"{col_min}{min_r}:{col_max}{novo_max_r}"

    if tabela.autoFilter:
        tabela.autoFilter.ref = f"{col_min}{min_r}:{col_max}{min_r}"

    return novo_max_r, min_c, max_c


def copiar_formulas(ws, linha_destino, min_c, max_c):
    linha_origem = linha_destino - 1
    if linha_origem < 1:
        return

    for col in range(min_c, max_c + 1):
        origem = ws.cell(row=linha_origem, column=col)
        destino = ws.cell(row=linha_destino, column=col)

        if isinstance(origem.value, str) and origem.value.startswith("="):
            try:
                destino.value = Translator(
                    origem.value, origin=origem.coordinate
                ).translate_formula(destino.coordinate)
            except Exception:
                destino.value = origem.value


# ==================================================
# LEITURA DE HOSTNAMES
# ==================================================
def ler_hostnames(caminho):
    try:
        wb = load_workbook(caminho, read_only=True)
        aba = wb.active
        hostnames = {
            str(linha[0]).strip()
            for linha in aba.iter_rows(min_row=2, max_col=1, values_only=True)
            if linha and linha[0]
        }

        if not hostnames:
            abortar("Nenhum hostname encontrado.")

        return sorted(hostnames)
    except AutomacaoErro:
        raise
    except Exception as erro:
        abortar(f"Erro ao abrir arquivo de hostnames: {erro}")
    finally:
        if "wb" in locals():
            wb.close()


# ==================================================
# INSERÇÕES NO EXCEL
# ==================================================
def inserir_nova_linha(aba, nome_tabela):
    linha, min_c, max_c = expandir_tabela(aba, nome_tabela)
    copiar_formulas(aba, linha, min_c, max_c)
    return linha


def inserir_chamado_pai(aba, requisicao, wo, descricao):
    linha = inserir_nova_linha(aba, NOME_TABELA_PAI)
    aba.cell(linha, PAI_COL_NUMERO, requisicao)
    aba.cell(linha, PAI_COL_WO, wo)
    aba.cell(linha, PAI_COL_DESCRICAO, descricao)


def inserir_atuacoes(aba, wo, descricao, hostnames):
    estilo_centro = Alignment(horizontal="center")
    estilo_negrito = Font(bold=True)

    for host in hostnames:
        linha = inserir_nova_linha(aba, NOME_TABELA_ATUACAO)

        cel_wo = aba.cell(linha, ATU_COL_CHAMADO_PAI, wo)
        cel_wo.alignment = estilo_centro

        cel_host = aba.cell(linha, ATU_COL_HOSTNAME, host.upper())
        cel_host.font = estilo_negrito

        aba.cell(linha, ATU_COL_DESCRICAO, descricao)
        aba.cell(linha, ATU_COL_STATUS, "Pendente")


def gerenciar_backup(arquivo_original, pasta_backup, limite=5):
    os.makedirs(pasta_backup, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_base = os.path.splitext(os.path.basename(arquivo_original))[0]

    shutil.copy2(
        arquivo_original,
        os.path.join(pasta_backup, f"{nome_base}_backup_{timestamp}.xlsx"),
    )

    backups = sorted(
        [
            os.path.join(pasta_backup, f)
            for f in os.listdir(pasta_backup)
            if f.startswith(f"{nome_base}_backup_") and f.lower().endswith(".xlsx")
        ],
        key=os.path.getmtime,
        reverse=True,
    )

    for old_path in backups[limite:]:
        try:
            os.remove(old_path)
        except OSError:
            pass


# ==================================================
# PROGRAMA PRINCIPAL
# ==================================================
def executar_automacao(requisicao, wo, software, cve, arquivo_hosts, arquivo_principal):
    validar_campos(requisicao, wo, software, cve, arquivo_hosts, arquivo_principal)
    validar_extensao_excel(arquivo_hosts)

    descricao = gerar_descricao(software, cve)

    validar_arquivo(arquivo_principal, "Arquivo principal")
    validar_arquivo(arquivo_hosts, "Arquivo de hostnames")

    try:
        wb = load_workbook(arquivo_principal)
    except PermissionError:
        abortar("Feche o Excel antes de executar o script.")
    except Exception as erro:
        abortar(f"Erro ao abrir planilha principal: {erro}")

    try:
        aba_pai = wb["🔎 CHAMADO PAI"]
        aba_atuacao = wb["📈 RESCALDOS- ATUALIZAÇÃO"]

        if any(
            str(linha[PAI_COL_WO - 1]).strip() == wo
            for linha in aba_pai.iter_rows(
                min_row=2, max_col=PAI_COL_WO, values_only=True
            )
        ):
            abortar("WO já existe na planilha.")

        hostnames = ler_hostnames(arquivo_hosts)

        inserir_chamado_pai(aba_pai, requisicao, wo, descricao)
        inserir_atuacoes(aba_atuacao, wo, descricao, hostnames)

        base_dir = os.path.dirname(arquivo_principal)
        gerenciar_backup(arquivo_principal, os.path.join(base_dir, "backup"))

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_name = tmp.name
        wb.save(tmp_name)
    finally:
        if "wb" in locals():
            wb.close()

    try:
        shutil.move(tmp_name, arquivo_principal)
    except Exception as e:
        abortar(f"Erro ao salvar o arquivo principal: {e}")

    return len(hostnames)
