import os

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def gerar_amostras():
    base_dir = os.path.join(os.path.dirname(__file__), "amostras")
    os.makedirs(base_dir, exist_ok=True)

    # 1. Planilha Principal
    wb_principal = Workbook()

    # Aba Chamado Pai
    ws_pai = wb_principal.active
    ws_pai.title = "🔎 CHAMADO PAI"
    ws_pai.append(["REQUISIÇÃO", "WO", "DESCRIÇÃO"])
    ws_pai.append(["REQ0000001", "WO0000001", "Exemplo de Requisição Anterior"])

    tab_pai = Table(displayName="Chamado_Pai", ref="A1:C2")
    style_pai = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tab_pai.tableStyleInfo = style_pai
    ws_pai.add_table(tab_pai)

    # Aba Atuação
    ws_atuacao = wb_principal.create_sheet(title="📈 RESCALDOS- ATUALIZAÇÃO")
    # A estrutura esperada pela automação:
    # Col 1: Chamado Pai, Col 2: Hostname, Col 4: Descrição, Col 5: Status
    ws_atuacao.append(["CHAMADO PAI", "HOSTNAME", "AMBIENTE", "DESCRIÇÃO", "STATUS"])
    ws_atuacao.append(
        ["WO0000001", "SRV-OLD-01", "PROD", "Atualização Anterior", "Concluído"]
    )

    tab_atuacao = Table(displayName="Atuação", ref="A1:E2")
    style_atuacao = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tab_atuacao.tableStyleInfo = style_atuacao
    ws_atuacao.add_table(tab_atuacao)

    caminho_principal = os.path.join(base_dir, "planilha_principal_exemplo.xlsx")
    wb_principal.save(caminho_principal)

    # 2. Arquivo de Máquinas
    wb_hosts = Workbook()
    ws_hosts = wb_hosts.active
    ws_hosts.title = "Máquinas"
    ws_hosts.append(["Hostname"])
    ws_hosts.append(["SRV-APP-01"])
    ws_hosts.append(["SRV-APP-02"])
    ws_hosts.append(["SRV-BD-01"])

    caminho_hosts = os.path.join(base_dir, "maquinas_exemplo.xlsx")
    wb_hosts.save(caminho_hosts)

    print(f"Arquivos de amostra gerados com sucesso na pasta '{base_dir}'!")


if __name__ == "__main__":
    gerar_amostras()
